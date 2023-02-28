import tempfile
import subprocess
import os
import os.path
import shutil
import sys
import re
#subprocess.Popen(f'explorer', )

def get_factures(factures:list[str]): 
    for facture in factures:
        for year in ('2022', '2023'):
            for mes in ("ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO", "JULIO", "AGOSTO", "SETIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"):
                print(f"Buscando en el mes de {mes} de {year}")
                path = f"G:\\1.USUARIOS\\FACTURACION\\COMPROBANTES\\{year}\\{mes}"
                lista_factures = os.listdir(path)
                for factura in lista_factures:
                    if re.search(facture, factura):
                        print(f"Comprobante: {facture} encontrado")
                        if get_facture(factura, mes):
                            return

def get_filename(codigo_referencia :str) -> str:
    index_split = 4
    serie = codigo_referencia[:index_split]
    numero = codigo_referencia[index_split:]
    nuevo_codigo = f"{serie}-0{numero}"
    return nuevo_codigo

def get_facture(facture, mes):
    facture_path = f"G:\\1.USUARIOS\\FACTURACION\\COMPROBANTES\\2022\\{mes}\\{facture}"
    
    copy_path = os.path.join("C:\\Users\\abernabel\\Desktop\\temp", facture)
    if not os.path.isdir("C:/Users/abernabel/Desktop/temp"):
        os.mkdir("C:/Users/abernabel/Desktop/temp")
    
    #print(f"copiando {facture_path} a {copy_path}")
    shutil.copyfile(facture_path, copy_path)
    return True
    #shutil.copyfile(facture_path, os.path.join(copy_path, facture.replace(".pdf",".XML")))
    
def process_input(argumento):
    if os.path.splitext(argumento)[1] == ".xlsx":
        import openpyxl
        wb_obj = openpyxl.load_workbook("input.xlsx")
 

        sheet_obj = wb_obj.active
        lista_cods = []
        for i in range(1, sheet_obj.max_row):

            lista_cods.append(get_filename(sheet_obj.cell(row = i, column = 1).value))
        
        return lista_cods
    else:
        lista_cods = argumento.split(",")
        lista_result = []
        for cod in lista_cods:
            lista_result.append(get_filename(cod))
        return lista_result

if __name__ == "__main__":
    
    entrada = sys.argv[1]
	
    factures = process_input(entrada)
    print(*factures, sep='\n')
    #pausa = input("nueva_data")
    get_factures(factures)

 	
